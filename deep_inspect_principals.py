import os
import openpyxl

files = {
    "Dulux": r"C:\Users\jamil\OneDrive\Documents\Reporting Prinsiple\Dulux\Copy of Template Report Untuk Sadata.xlsx",
    "Fonterra": r"C:\Users\jamil\OneDrive\Documents\Reporting Prinsiple\Fonterra\Copy of All Report Fonterra.xlsx",
    "Mamasuka": r"C:\Users\jamil\OneDrive\Documents\Reporting Prinsiple\Mamasuka\Copy of RAW DATA - REPORTING  ATTANDANCE MAMASUKA.xlsx"
}

for name, path in files.items():
    print("="*70)
    print(f"PRINCIPAL: {name.upper()}")
    print("="*70)
    if not os.path.exists(path):
        print("File not found!")
        continue
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    print("Sheet Names:", wb.sheetnames)
    for sname in wb.sheetnames:
        ws = wb[sname]
        headers = []
        for row in ws.iter_rows(values_only=True):
            if any(row):
                headers = [str(c).strip() for c in row if c is not None]
                break
        print(f"\n  [Sheet: {sname}] (Columns count: {len(headers)})")
        print(f"   Fields: {', '.join(headers[:20])}")
        if len(headers) > 20:
            print(f"   ... dan {len(headers)-20} kolom lainnya: {', '.join(headers[20:35])}")

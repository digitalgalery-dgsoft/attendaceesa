import openpyxl
import sqlite3
import gzip
import shutil
import os
import sys
import re

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

f2026 = r"C:\Users\jamil\Downloads\Data Dulux\Daily Maintenance\All POST Daily Maintenance Jan-Jul 2026.xlsx"
f2025 = r"C:\Users\jamil\Downloads\Data Dulux\Daily Maintenance\All POST Daily Maintenance Jan-Des 2025.xlsx"

out_dir = r"g:\My File\Project APlikasi Absensi\New\att-admin-v12\storage\app\dulux_data"
os.makedirs(out_dir, exist_ok=True)
db_path = os.path.join(out_dir, "daily_maintenance.sqlite")
gz_path = os.path.join(out_dir, "daily_maintenance.sqlite.gz")

if os.path.exists(db_path):
    os.remove(db_path)

conn = sqlite3.connect(db_path)
cur = conn.cursor()

# Create table
cur.execute("""
CREATE TABLE dm_raw (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    year INTEGER,
    month INTEGER,
    submission_date TEXT,
    tanggal_report TEXT,
    store_name TEXT,
    sap_code TEXT,
    category TEXT,
    rsm_area TEXT,
    area TEXT,
    tl_name TEXT,
    machine_type TEXT,
    machine_no TEXT,
    dc_name TEXT,
    kesimpulan TEXT,
    tinta_ok INTEGER DEFAULT 0,
    d200_nozzle_ok INTEGER DEFAULT 0,
    d200_cup_ok INTEGER DEFAULT 0,
    discovery_brush_ok INTEGER DEFAULT 0,
    manual_nozzle_ok INTEGER DEFAULT 0,
    mix2win_steps_ok INTEGER DEFAULT 0,
    pembersihan_all_ok INTEGER DEFAULT 0,
    is_compliant INTEGER DEFAULT 1
);
""")

cur.execute("CREATE INDEX idx_dm_year_month ON dm_raw(year, month);")
cur.execute("CREATE INDEX idx_dm_rsm ON dm_raw(rsm_area);")
cur.execute("CREATE INDEX idx_dm_area ON dm_raw(area);")
cur.execute("CREATE INDEX idx_dm_store ON dm_raw(store_name);")
cur.execute("CREATE INDEX idx_dm_sap ON dm_raw(sap_code);")
cur.execute("CREATE INDEX idx_dm_mach_type ON dm_raw(machine_type);")
cur.execute("CREATE INDEX idx_dm_cat ON dm_raw(category);")

month_map = {
    'jan': 1, 'feb': 2, 'mar': 3, 'maret': 3, 'apr': 4, 'april': 4,
    'mei': 5, 'may': 5, 'jun': 6, 'juni': 6, 'jul': 7, 'juli': 7,
    'agu': 8, 'agustus': 8, 'agt': 8, 'aug': 8, 'sep': 9, 'sept': 9,
    'okt': 10, 'oktober': 10, 'oct': 10, 'nov': 11, 'des': 12, 'dec': 12
}

def clean_val(v):
    if v is None:
        return ''
    return str(v).strip()

def is_date(s):
    if not s:
        return False
    s = str(s).strip()
    return bool(re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', s))

def process_file(fpath, year_fixed):
    print(f"\nProcessing {fpath} (Year {year_fixed})...")
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    
    total_file_rows = 0
    
    for sname in wb.sheetnames:
        m_int = month_map.get(sname.lower().strip()[:3], 1)
        ws = wb[sname]
        header_row = None
        batch = []
        
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            
            # Check if this row is a header row
            if 'Submission Date' in str(row[0]) or 'Nama Toko' in str(row[1] if len(row) > 1 else ''):
                header_row = [clean_val(c) for c in row]
                continue
            
            # Skip if we haven't seen a header or row is empty
            if not any(c is not None for c in row):
                continue
            
            sub_date = clean_val(row[0])
            store_name = clean_val(row[1]) if len(row) > 1 else ''
            area = clean_val(row[2]) if len(row) > 2 else ''
            rsm = clean_val(row[3]) if len(row) > 3 else ''
            sap = clean_val(row[4]) if len(row) > 4 else ''
            cat = clean_val(row[5]) if len(row) > 5 else ''
            
            c6 = clean_val(row[6]) if len(row) > 6 else ''
            c7 = clean_val(row[7]) if len(row) > 7 else ''
            c8 = clean_val(row[8]) if len(row) > 8 else ''
            c9 = clean_val(row[9]) if len(row) > 9 else ''
            c10 = clean_val(row[10]) if len(row) > 10 else ''
            c11 = clean_val(row[11]) if len(row) > 11 else ''
            
            # Check if double TL (c10 is date or c6 == c7)
            if is_date(c10) or (c6 and c7 and c6.lower() == c7.lower()):
                tl = c6
                mtype = c8
                mno = c9
                tgl = c10
                dc = c11
                tinta_val = clean_val(row[12]) if len(row) > 12 else ''
                kesimpulan = clean_val(row[15]) if len(row) > 15 else ''
                chk_offset = 1
            else:
                tl = c6
                mtype = c7
                mno = c8
                tgl = c9
                dc = c10
                tinta_val = clean_val(row[11]) if len(row) > 11 else ''
                kesimpulan = clean_val(row[14]) if len(row) > 14 else ''
                chk_offset = 0
            
            # If tgl is empty but sub_date is present, extract date part
            if not tgl and sub_date:
                tgl = sub_date.split(' ')[0]
            
            # Checklists
            tinta_ok = 1 if tinta_val.upper() in ['YA', 'OK', '1', 'TRUE'] else 0
            
            # D200 nozzle & cup (standard: col 15, col 18)
            d200_noz_idx = 15 + chk_offset
            d200_cup_idx = 18 + chk_offset
            d200_noz = 1 if len(row) > d200_noz_idx and clean_val(row[d200_noz_idx]).upper() in ['YA', 'OK', '1', 'TRUE'] else 0
            d200_cup = 1 if len(row) > d200_cup_idx and clean_val(row[d200_cup_idx]).upper() in ['YA', 'OK', '1', 'TRUE'] else 0
            
            # Discovery brush (standard: col 27)
            disc_br_idx = 27 + chk_offset
            disc_br = 1 if len(row) > disc_br_idx and clean_val(row[disc_br_idx]).upper() in ['YA', 'OK', '1', 'TRUE'] else 0
            
            # Manual nozzle (standard: col 33)
            man_noz_idx = 33 + chk_offset
            man_noz = 1 if len(row) > man_noz_idx and clean_val(row[man_noz_idx]).upper() in ['YA', 'OK', '1', 'TRUE'] else 0
            
            # Mix2Win steps (steps 1 to 12)
            mix_steps = 0
            mix_start_idx = 39 + chk_offset
            for step_i in range(12):
                col_i = mix_start_idx + (step_i * 3)
                if len(row) > col_i and clean_val(row[col_i]).upper() in ['YA', 'OK', '1', 'TRUE']:
                    mix_steps += 1
            
            # Pembersihan (4 checklist items: col 75, 78, 81, 84)
            pembersihan_cnt = 0
            pem_start_idx = 75 + chk_offset
            for pem_i in range(4):
                col_i = pem_start_idx + (pem_i * 3)
                if len(row) > col_i and clean_val(row[col_i]).upper() in ['YA', 'OK', '1', 'TRUE']:
                    pembersihan_cnt += 1
            pembersihan_all = 1 if pembersihan_cnt >= 3 else 0
            
            # Normalize SAP
            if not sap and '-' in store_name:
                parts = store_name.split('-')
                last_part = parts[-1].strip()
                if last_part.isdigit():
                    sap = last_part
            
            # Standardize RSM
            rsm_clean = rsm.strip()
            if rsm_clean.upper() in ['BALI NUSRA PUMA', 'BALI NUSRA']:
                rsm_clean = 'Bali Nusra'
            elif rsm_clean.upper() in ['EAST JAVA', 'JAWA TIMUR']:
                rsm_clean = 'East Java'
            elif rsm_clean.upper() in ['WEST JAVA', 'JAWA BARAT']:
                rsm_clean = 'West Java'
            elif rsm_clean.upper() in ['CENTRAL SUMATERA', 'CENTAL SUMATERA', 'SUMATERA TENGAH']:
                rsm_clean = 'Central Sumatera'
            elif rsm_clean.upper() in ['SUMATERA UTARA', 'NORTH SUMATERA']:
                rsm_clean = 'North Sumatera'
            elif rsm_clean.upper() in ['SUMATERA SELATAN', 'SOUTH SUMATERA']:
                rsm_clean = 'South Sumatera'
            elif rsm_clean.upper() in ['LSO', 'LSO EAST', 'LSO WEST']:
                rsm_clean = 'LSO'
            elif rsm_clean.upper() in ['SULAWESI']:
                rsm_clean = 'Sulawesi'
            elif rsm_clean.upper() in ['KALIMANTAN']:
                rsm_clean = 'Kalimantan'
            elif rsm_clean.upper() in ['GREATER JAKARTA']:
                rsm_clean = 'Greater Jakarta'
            elif rsm_clean.upper() in ['JAWA TENGAH', 'NORTH CENTRAL JAVA', 'SOUTH CENTRAL JAVA']:
                rsm_clean = 'Central Java'
            
            # Normalize Machine Type
            comb_mach = (mtype + " " + mno).upper()
            if 'DISCOVERY' in comb_mach or 'DISCV' in comb_mach:
                mtype_clean = 'Discovery'
            elif 'X-SMART' in comb_mach or 'XSMART' in comb_mach or 'X SMART' in comb_mach or '670000001' in comb_mach:
                mtype_clean = 'X-Smart'
            elif 'XPROTINT' in comb_mach or 'X-PROTINT' in comb_mach or '720000010' in comb_mach or 'PROTINT' in comb_mach:
                mtype_clean = 'Xprotint'
            elif 'D200' in comb_mach or 'D10B' in comb_mach or 'D14B' in comb_mach or 'D11B' in comb_mach or 'D8B' in comb_mach:
                mtype_clean = 'D200'
            elif 'MANUAL' in comb_mach:
                mtype_clean = 'Manual'
            elif 'FAST' in comb_mach or 'FLUID' in comb_mach:
                mtype_clean = 'Fast & Fluid'
            else:
                mtype_clean = 'Other'
            
            batch.append((
                year_fixed, m_int, sub_date, tgl, store_name, sap, cat, rsm_clean, area,
                tl, mtype_clean, mno, dc, kesimpulan,
                tinta_ok, d200_noz, d200_cup, disc_br, man_noz, mix_steps, pembersihan_all, 1
            ))
            
            if len(batch) >= 2000:
                cur.executemany("""
                INSERT INTO dm_raw (
                    year, month, submission_date, tanggal_report, store_name, sap_code, category, rsm_area, area,
                    tl_name, machine_type, machine_no, dc_name, kesimpulan,
                    tinta_ok, d200_nozzle_ok, d200_cup_ok, discovery_brush_ok, manual_nozzle_ok,
                    mix2win_steps_ok, pembersihan_all_ok, is_compliant
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, batch)
                conn.commit()
                total_file_rows += len(batch)
                batch = []
        
        if batch:
            cur.executemany("""
            INSERT INTO dm_raw (
                year, month, submission_date, tanggal_report, store_name, sap_code, category, rsm_area, area,
                tl_name, machine_type, machine_no, dc_name, kesimpulan,
                tinta_ok, d200_nozzle_ok, d200_cup_ok, discovery_brush_ok, manual_nozzle_ok,
                mix2win_steps_ok, pembersihan_all_ok, is_compliant
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, batch)
            conn.commit()
            total_file_rows += len(batch)
            batch = []
            
        print(f" - Completed Sheet [{sname}]: Month {m_int}")
    
    print(f"Total inserted from {fpath}: {total_file_rows} rows.")

process_file(f2025, 2025)
process_file(f2026, 2026)

# Verify count
cur.execute("SELECT COUNT(*) FROM dm_raw")
total_rows = cur.fetchone()[0]
print(f"\n=======================================================")
print(f"TOTAL ROWS IN SQLITE: {total_rows}")
print(f"=======================================================")

# Verify anomalies in machine_no
cur.execute("SELECT COUNT(*) FROM dm_raw WHERE machine_no LIKE '%/2026%' OR machine_no LIKE '%/2025%'")
date_in_mno_cnt = cur.fetchone()[0]
print(f"Total rows with date in machine_no: {date_in_mno_cnt}")

conn.close()

# Create .gz archive
print(f"\nCompressing to {gz_path}...")
with open(db_path, 'rb') as f_in:
    with gzip.open(gz_path, 'wb', compresslevel=9) as f_out:
        shutil.copyfileobj(f_in, f_out)

db_size_mb = os.path.getsize(db_path) / (1024 * 1024)
gz_size_mb = os.path.getsize(gz_path) / (1024 * 1024)
print(f"Done! SQLite DB Size: {db_size_mb:.2f} MB, GZ Archive Size: {gz_size_mb:.2f} MB")

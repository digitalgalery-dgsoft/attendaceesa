import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')
f2026 = r"C:\Users\jamil\Downloads\Data Dulux\Daily Maintenance\All POST Daily Maintenance Jan-Jul 2026.xlsx"
wb = openpyxl.load_workbook(f2026, read_only=True, data_only=True)

for sname in wb.sheetnames:
    ws = wb[sname]
    for row in ws.iter_rows(values_only=True):
        if any(c in ['Submission Date', 'Nama Toko'] for c in row if c):
            headers = [str(c).strip() if c is not None else '' for c in row]
            print(f"\n==================== Sheet [{sname}] ({len(headers)} cols) ====================")
            for idx, h in enumerate(headers[:20]):
                if h:
                    print(f"  Col {idx:2d}: {h}")
            break

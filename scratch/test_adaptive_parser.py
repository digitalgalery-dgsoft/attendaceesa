import openpyxl
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

f2026 = r"C:\Users\jamil\Downloads\Data Dulux\Daily Maintenance\All POST Daily Maintenance Jan-Jul 2026.xlsx"
wb = openpyxl.load_workbook(f2026, read_only=True, data_only=True)
ws = wb['Juni']

def is_date(s):
    if not s:
        return False
    s = str(s).strip()
    return bool(re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', s))

def clean_val(v):
    if v is None:
        return ''
    return str(v).strip()

row_count = 0
date_in_machine_no_count = 0

for row in ws.iter_rows(values_only=True):
    if not row or not row[0] or 'Submission Date' in str(row[0]):
        continue
    row_count += 1
    
    # Adaptive detection of row layout
    # Col 0: Submission Date, 1: Store, 2: Area, 3: RSM, 4: SAP, 5: Cat
    sub_date = clean_val(row[0])
    store_name = clean_val(row[1])
    area = clean_val(row[2])
    rsm = clean_val(row[3])
    sap = clean_val(row[4])
    cat = clean_val(row[5])
    
    c6 = clean_val(row[6]) if len(row) > 6 else ''
    c7 = clean_val(row[7]) if len(row) > 7 else ''
    c8 = clean_val(row[8]) if len(row) > 8 else ''
    c9 = clean_val(row[9]) if len(row) > 9 else ''
    c10 = clean_val(row[10]) if len(row) > 10 else ''
    c11 = clean_val(row[11]) if len(row) > 11 else ''
    
    # Check if double TL (c10 is date or c6 == c7)
    if is_date(c10) or (c6 and c7 and c6.lower() == c7.lower()):
        # Double TL layout
        tl = c6
        mtype = c8
        mno = c9
        tgl = c10
        dc = c11
        tinta_val = clean_val(row[12]) if len(row) > 12 else ''
        kesimpulan = clean_val(row[15]) if len(row) > 15 else ''
    else:
        # Standard layout (1 TL)
        tl = c6
        mtype = c7
        mno = c8
        tgl = c9
        dc = c10
        tinta_val = clean_val(row[11]) if len(row) > 11 else ''
        kesimpulan = clean_val(row[14]) if len(row) > 14 else ''
    
    if is_date(mno):
        date_in_machine_no_count += 1
        if date_in_machine_no_count <= 5:
            print(f"ANOMALY at row {row_count}: mno='{mno}', tgl='{tgl}', dc='{dc}'")

print(f"\nTotal rows processed in Sheet [Juni]: {row_count}")
print(f"Total rows where machine_no contains a date: {date_in_machine_no_count}")

# Check row 1104 (the one that previously failed)
print("\nSample check around row 1100-1105 in June:")
row_idx = 0
for row in ws.iter_rows(values_only=True):
    if not row or not row[0] or 'Submission Date' in str(row[0]):
        continue
    row_idx += 1
    if 1100 <= row_idx <= 1106:
        c6 = clean_val(row[6]) if len(row) > 6 else ''
        c7 = clean_val(row[7]) if len(row) > 7 else ''
        c8 = clean_val(row[8]) if len(row) > 8 else ''
        c9 = clean_val(row[9]) if len(row) > 9 else ''
        c10 = clean_val(row[10]) if len(row) > 10 else ''
        c11 = clean_val(row[11]) if len(row) > 11 else ''
        
        if is_date(c10) or (c6 and c7 and c6.lower() == c7.lower()):
            tl, mtype, mno, tgl, dc = c6, c8, c9, c10, c11
        else:
            tl, mtype, mno, tgl, dc = c6, c7, c8, c9, c10
        print(f" Row {row_idx}: Store='{row[1]}', TL='{tl}', MType='{mtype}', MNo='{mno}', Tgl='{tgl}', DC='{dc}'")

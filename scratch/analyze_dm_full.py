import openpyxl
import os

f2026 = r"C:\Users\jamil\Downloads\Data Dulux\Daily Maintenance\All POST Daily Maintenance Jan-Jul 2026.xlsx"
f2025 = r"C:\Users\jamil\Downloads\Data Dulux\Daily Maintenance\All POST Daily Maintenance Jan-Des 2025.xlsx"

def analyze_wb(fpath, label):
    print("=" * 70)
    print(f"ANALYZING: {label} ({fpath})")
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    total_all = 0
    all_headers = {}
    sample_machine_types = set()
    sample_categories = set()
    sample_areas = set()
    sample_rsm = set()
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        header_row = None
        count = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                pass
            if any(cell == 'Submission Date' or cell == 'Nama Toko' for cell in row if cell):
                header_row = [str(c).strip() if c is not None else f"Col_{i}" for i, c in enumerate(row)]
                continue
            if header_row and any(c is not None for c in row):
                count += 1
                # Sample some columns
                # Let's see if we can find Tipe Mesin, Kategori, Area
                for col_name, val in zip(header_row, row):
                    if val is not None:
                        if 'Tipe Mesin' in col_name:
                            sample_machine_types.add(str(val).strip())
                        if 'Kategori' in col_name:
                            sample_categories.add(str(val).strip())
                        if col_name == 'Area':
                            sample_areas.add(str(val).strip())
                        if 'RSM' in col_name:
                            sample_rsm.add(str(val).strip())
        
        print(f" - Sheet [{sname}]: {count} rows | Headers ({len(header_row) if header_row else 0} cols)")
        if header_row:
            all_headers[sname] = header_row
        total_all += count
    
    wb.close()
    print(f"\nTOTAL ROWS IN {label}: {total_all}")
    print(f"Machine Types: {sample_machine_types}")
    print(f"Categories: {sample_categories}")
    print(f"RSM Areas: {sample_rsm}")
    print(f"Areas count: {len(sample_areas)}")
    return all_headers

headers_2026 = analyze_wb(f2026, "2026 (Jan-Jul)")
headers_2025 = analyze_wb(f2025, "2025 (Jan-Des)")

# Compare header columns
print("\n" + "=" * 70)
print("HEADER COMPARISON:")
if 'Juli' in headers_2026:
    print("\n2026 Juli Headers (Total cols: " + str(len(headers_2026['Juli'])) + "):")
    for i, h in enumerate(headers_2026['Juli']):
        print(f"  {i+1}. {h}")

if 'Juli' in headers_2025:
    print("\n2025 Juli Headers (Total cols: " + str(len(headers_2025['Juli'])) + "):")
    for i, h in enumerate(headers_2025['Juli']):
        print(f"  {i+1}. {h}")

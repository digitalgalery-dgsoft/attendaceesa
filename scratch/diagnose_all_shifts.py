import openpyxl
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

f2026 = r"C:\Users\jamil\Downloads\Data Dulux\Daily Maintenance\All POST Daily Maintenance Jan-Jul 2026.xlsx"
f2025 = r"C:\Users\jamil\Downloads\Data Dulux\Daily Maintenance\All POST Daily Maintenance Jan-Des 2025.xlsx"

def is_date(s):
    if not s:
        return False
    s = str(s).strip()
    return bool(re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', s))

def is_known_mtype(s):
    if not s:
        return False
    s = str(s).upper()
    return any(k in s for k in ['D200', 'DISCOVERY', 'X-SMART', 'XSMART', 'XPROTINT', 'MANUAL', 'FAST'])

for fpath, yr in [(f2026, 2026), (f2025, 2025)]:
    print(f"\n==================== CHECKING {yr} ====================")
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        row_idx = 0
        shifted_count = 0
        for row in ws.iter_rows(values_only=True):
            row_idx += 1
            if not row or not row[0] or 'Submission Date' in str(row[0]):
                continue
            
            # Check where the date column is
            # In standard 89-col layout:
            # 0: Sub Date, 1: Store, 2: Area, 3: RSM, 4: SAP, 5: Cat, 6: TL, 7: MType, 8: MNo, 9: Date, 10: DC
            # In 90-col double-TL layout:
            # 0: Sub Date, 1: Store, 2: Area, 3: RSM, 4: SAP, 5: Cat, 6: TL, 7: TL, 8: MType, 9: MNo, 10: Date, 11: DC
            
            # Let's inspect Col 7, 8, 9, 10
            c6 = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ''
            c7 = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ''
            c8 = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ''
            c9 = str(row[9]).strip() if len(row) > 9 and row[9] is not None else ''
            c10 = str(row[10]).strip() if len(row) > 10 and row[10] is not None else ''
            
            # If c9 is date and c7 is mtype -> standard layout (1 TL)
            # If c10 is date and c8 is mtype -> double TL layout
            if is_date(c9) and (is_known_mtype(c7) or c6 == c7 or not is_date(c10)):
                pass # standard layout
            elif is_date(c10) and is_known_mtype(c8):
                pass # double TL layout
            else:
                shifted_count += 1
                if shifted_count <= 5:
                    print(f"[{sname} Row {row_idx}] 6:'{c6}' | 7:'{c7}' | 8:'{c8}' | 9:'{c9}' | 10:'{c10}'")
        
        print(f"Sheet [{sname}]: Total rows checked: {row_idx}, Anomaly rows: {shifted_count}")

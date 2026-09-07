import openpyxl

f2026 = r"C:\Users\jamil\Downloads\Data Dulux\Daily Maintenance\All POST Daily Maintenance Jan-Jul 2026.xlsx"
wb = openpyxl.load_workbook(f2026, read_only=True, data_only=True)
sheet_names = wb.sheetnames
print("2026 Sheet Names:", sheet_names)

# Inspect 'Juli' and 'Jan'
for sname in ['Jan', 'Juli']:
    if sname in sheet_names:
        ws = wb[sname]
        print(f"\n--- 2026 Sheet [{sname}] ---")
        rows = list(ws.iter_rows(values_only=True, max_row=15))
        for idx, r in enumerate(rows):
            # filter non None
            non_empty = [c for c in r if c is not None]
            if len(non_empty) > 0:
                print(f"Row {idx+1}: {r[:20]}")
                if idx > 8:
                    break

wb.close()

f2025 = r"C:\Users\jamil\Downloads\Data Dulux\Daily Maintenance\All POST Daily Maintenance Jan-Des 2025.xlsx"
wb25 = openpyxl.load_workbook(f2025, read_only=True, data_only=True)
print("\n2025 Sheet Names:", wb25.sheetnames)
ws25 = wb25['Juli']
print(f"\n--- 2025 Sheet [Juli] ---")
rows25 = list(ws25.iter_rows(values_only=True, max_row=15))
for idx, r in enumerate(rows25):
    non_empty = [c for c in r if c is not None]
    if len(non_empty) > 0:
        print(f"Row {idx+1}: {r[:20]}")
        if idx > 8:
            break
wb25.close()

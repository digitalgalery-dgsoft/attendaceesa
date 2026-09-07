import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

f2026 = r"C:\Users\jamil\Downloads\Data Dulux\Daily Maintenance\All POST Daily Maintenance Jan-Jul 2026.xlsx"
wb = openpyxl.load_workbook(f2026, read_only=True, data_only=True)

def clean_val(v):
    return str(v).strip() if v is not None else ''

for sname in wb.sheetnames:
    ws = wb[sname]
    for row in ws.iter_rows(values_only=True):
        if any(c in ['Submission Date', 'Nama Toko'] for c in row if c):
            header_row = [clean_val(c) for c in row]
            col_map = {}
            for i, h in enumerate(header_row):
                h_lower = h.lower()
                if 'submission date' in h_lower:
                    col_map['submission_date'] = i
                elif 'nama toko' in h_lower:
                    col_map['store_name'] = i
                elif 'area' == h_lower:
                    col_map['area'] = i
                elif 'rsm area' in h_lower:
                    col_map['rsm_area'] = i
                elif 'kode sap' in h_lower or 'sap member' in h_lower:
                    col_map['sap_code'] = i
                elif 'kategori toko' in h_lower or 'kategori' == h_lower:
                    col_map['category'] = i
                elif 'nama tl' in h_lower or 'tl' == h_lower:
                    col_map['tl_name'] = i
                elif 'tipe mesin' in h_lower:
                    col_map['machine_type'] = i
                elif 'no mesin' in h_lower or 'no seri' in h_lower:
                    col_map['machine_no'] = i
                elif 'tanggal' == h_lower:
                    col_map['tanggal'] = i
                elif 'nama dc' in h_lower or 'promoter' in h_lower:
                    col_map['dc_name'] = i
                elif 'kesimpulan' in h_lower:
                    col_map['kesimpulan'] = i
            
            print(f"\nSheet [{sname}] col_map:")
            for k, v in col_map.items():
                print(f"  {k:15s} => col {v:2d} ('{header_row[v]}')")
            
            # Print first data row
            for drow in ws.iter_rows(values_only=True):
                if drow[0] != 'Submission Date' and drow[0] is not None:
                    print("  Sample data row:")
                    for k, v in col_map.items():
                        val = clean_val(drow[v]) if v < len(drow) else ''
                        print(f"    {k:15s} ({v}): {val}")
                    break
            break

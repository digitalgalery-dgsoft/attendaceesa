import openpyxl
import sqlite3
import os

excel_path = r'C:\Users\jamil\Downloads\Data Dulux\Stock 2026 (1)\Stock End 2026\202607 Stock End Juli 2026 AMK AKP Updt CKDB and ACE.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws_summ = wb['Summ']

# 2025 Months in Summ:
# Col 19: JANUARI (DULUX: 19, CATYLAC: 20, TOTAL: 21)
# Col 29: FEBRUARI (DULUX: 29, CATYLAC: 30, TOTAL: 31)
# Col 39: MARET (DULUX: 39, CATYLAC: 40, TOTAL: 41)
# Col 49: APRIL (DULUX: 49, CATYLAC: 50, TOTAL: 51)
# Col 59: MAY (DULUX: 59, CATYLAC: 60, TOTAL: 61)
# Col 69: JUNE (DULUX: 69, CATYLAC: 70, TOTAL: 71)
# Col 79: July (DULUX: 79, CATYLAC: 80, TOTAL: 81)

month_cols_2025 = [
    (1, 19, 20, 21),
    (2, 29, 30, 31),
    (3, 39, 40, 41),
    (4, 49, 50, 51),
    (5, 59, 60, 61),
    (6, 69, 70, 71),
    (7, 79, 80, 81),
]

# Let's see how many valid rows
tot_rows = 0
tot_dulux = 0
tot_catylac = 0

for r in range(6, ws_summ.max_row + 1):
    sap = ws_summ.cell(r, 4).value # SAP Member
    store_name = ws_summ.cell(r, 6).value
    if not sap and not store_name:
        continue
    tot_rows += 1
    for m, c_dulux, c_cat, c_tot in month_cols_2025:
        v_d = ws_summ.cell(r, c_dulux).value or 0
        v_c = ws_summ.cell(r, c_cat).value or 0
        if isinstance(v_d, (int, float)):
            tot_dulux += v_d
        if isinstance(v_c, (int, float)):
            tot_catylac += v_c

print(f'2025 Summ Store count: {tot_rows}')
print(f'2025 YTD (M1-M7) Dulux Stock: {tot_dulux:,.2f} L')
print(f'2025 YTD (M1-M7) Catylac Stock: {tot_catylac:,.2f} L')
print(f'2025 YTD (M1-M7) Total Stock: {tot_dulux + tot_catylac:,.2f} L')

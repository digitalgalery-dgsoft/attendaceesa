import openpyxl

excel_path = r'C:\Users\jamil\Downloads\Data Dulux\Stock 2026 (1)\Stock End 2026\202607 Stock End Juli 2026 AMK AKP Updt CKDB and ACE.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws_summ = wb['Summ']

print('=== SUMM ROW 4 MONTH HEADERS ===')
for c in range(18, ws_summ.max_column + 1):
    r4 = ws_summ.cell(4, c).value
    r5 = ws_summ.cell(5, c).value
    if r4 is not None:
        print(f'Col {c}: Month="{r4}", FirstHeader="{r5}"')

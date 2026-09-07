import openpyxl

excel_path = r'C:\Users\jamil\Downloads\Data Dulux\Stock 2026 (1)\Stock End 2026\202607 Stock End Juli 2026 AMK AKP Updt CKDB and ACE.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)

# 1. Inspect Summ sheet columns and structure
ws_summ = wb['Summ']
print('=== SUMM SHEET ANALYSIS ===')
print('Max row:', ws_summ.max_row, 'Max col:', ws_summ.max_column)

# Row 4 (Month labels) and Row 5 (Column headers)
col_info = []
current_month = ''
for c in range(1, ws_summ.max_column + 1):
    m_val = ws_summ.cell(4, c).value
    if m_val:
        current_month = str(m_val).strip()
    h_val = ws_summ.cell(5, c).value
    col_info.append((c, current_month, h_val))

print('\nFirst 20 base columns (Store meta):')
for c, m, h in col_info[:20]:
    print(f'  Col {c}: Month="{m}", Header="{h}"')

# Let's see monthly metric blocks in Summ
print('\nMonthly metric groups in Summ:')
seen_months = {}
for c, m, h in col_info[17:]:
    if m:
        if m not in seen_months:
            seen_months[m] = []
        seen_months[m].append((c, h))

for m, cols in seen_months.items():
    print(f'\nMonth "{m}" ({len(cols)} columns):')
    for c, h in cols:
        print(f'    Col {c}: {h}')

# 2. Inspect Pivotable sheet
ws_piv = wb['Pivotable']
print('\n=== PIVOTABLE SHEET ANALYSIS ===')
print('Max row:', ws_piv.max_row, 'Max col:', ws_piv.max_column)
for r in range(1, 10):
    row_vals = [ws_piv.cell(r, c).value for c in range(1, ws_piv.max_column + 1)]
    print(f'Row {r}:', row_vals)

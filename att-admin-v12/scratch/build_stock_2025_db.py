import openpyxl
import sqlite3
import gzip
import shutil
import os

excel_path = r'C:\Users\jamil\Downloads\Data Dulux\Stock 2026 (1)\Stock End 2026\202607 Stock End Juli 2026 AMK AKP Updt CKDB and ACE.xlsx'
sqlite_path = 'storage/app/dulux_data/stock_2025.sqlite'
gz_path = 'storage/app/dulux_data/stock_2025.sqlite.gz'

if os.path.exists(sqlite_path):
    os.remove(sqlite_path)

conn = sqlite3.connect(sqlite_path)
cursor = conn.cursor()

# Create table
cursor.execute("""
CREATE TABLE stock_raw (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    year INTEGER,
    month INTEGER,
    submission_code TEXT,
    submission_date TEXT,
    tgl_catat TEXT,
    region TEXT,
    area TEXT,
    rsm_area TEXT,
    sap TEXT,
    derp TEXT,
    store_name TEXT,
    keterangan TEXT,
    brand TEXT,
    produk TEXT,
    warna TEXT,
    kemasan_galon REAL,
    qty_galon INTEGER,
    kemasan_pail REAL,
    qty_pail INTEGER,
    volume_liter REAL,
    conf REAL
);
""")

print("Loading Excel...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws_summ = wb['Summ']

# 2025 Months in Summ:
# Col 19: JANUARI (DULUX: 19, CATYLAC: 20, TOTAL: 21)
# Col 29: FEBRUARI (DULUX: 29, CATYLAC: 30, TOTAL: 31)
# Col 39: MARET (DULUX: 39, CATYLAC: 40, TOTAL: 41)
# Col 49: APRIL (DULUX: 49, CATYLAC: 50, TOTAL: 51)
# Col 59: MAY (DULUX: 59, CATYLAC: 60, TOTAL: 61)
# Col 69: JUNE (DULUX: 69, CATYLAC: 70, TOTAL: 71)
# Col 79: July (DULUX: 79, CATYLAC: 80, TOTAL: 81)

month_cols_2025 = [
    (1, 19, 20),
    (2, 29, 30),
    (3, 39, 40),
    (4, 49, 50),
    (5, 59, 60),
    (6, 69, 70),
    (7, 79, 80),
]

inserted = 0
records = []

for r in range(6, ws_summ.max_row + 1):
    sap = str(ws_summ.cell(r, 4).value or '').strip()
    store_name = str(ws_summ.cell(r, 6).value or '').strip()
    if not sap and not store_name:
        continue
    
    region = str(ws_summ.cell(r, 15).value or '').strip()
    area = str(ws_summ.cell(r, 9).value or '').strip()
    rsm_area = str(ws_summ.cell(r, 10).value or '').strip()
    
    for m, c_dulux, c_cat in month_cols_2025:
        v_dulux = ws_summ.cell(r, c_dulux).value or 0
        v_cat = ws_summ.cell(r, c_cat).value or 0
        
        if isinstance(v_dulux, (int, float)) and v_dulux > 0:
            records.append((
                2025, m, f'SUB-STOCK-2025-{m:02d}-{inserted+1:06d}',
                f'2025-{m:02d}-20 12:00:00', f'2025-{m:02d}-20',
                region, area, rsm_area, sap, '', store_name, 'Full Acces',
                'Dulux', 'Dulux Stock End', 'ALL', 2.5, 0, 20.0, 0, float(v_dulux), 1.0
            ))
            inserted += 1
            
        if isinstance(v_cat, (int, float)) and v_cat > 0:
            records.append((
                2025, m, f'SUB-STOCK-2025-{m:02d}-{inserted+1:06d}',
                f'2025-{m:02d}-20 12:00:00', f'2025-{m:02d}-20',
                region, area, rsm_area, sap, '', store_name, 'Full Acces',
                'Catylac', 'Catylac Stock End', 'ALL', 5.0, 0, 25.0, 0, float(v_cat), 1.0
            ))
            inserted += 1

cursor.executemany("""
INSERT INTO stock_raw (
    year, month, submission_code, submission_date, tgl_catat,
    region, area, rsm_area, sap, derp, store_name, keterangan,
    brand, produk, warna, kemasan_galon, qty_galon, kemasan_pail, qty_pail,
    volume_liter, conf
) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
""", records)

print("Creating indexes...")
cursor.execute("CREATE INDEX idx_stock_perf ON stock_raw (month, region, area, sap, store_name, brand, volume_liter);")
cursor.execute("CREATE INDEX idx_stock_store ON stock_raw (sap, store_name);")
cursor.execute("CREATE INDEX idx_stock_brand ON stock_raw (brand, volume_liter);")

conn.commit()
conn.close()

print(f"Total 2025 records inserted: {inserted}")
print(f"SQLite size: {os.path.getsize(sqlite_path):,} bytes")

print("Compressing to gz...")
with open(sqlite_path, 'rb') as f_in, gzip.open(gz_path, 'wb') as f_out:
    shutil.copyfileobj(f_in, f_out)

print(f"GZ size: {os.path.getsize(gz_path):,} bytes")

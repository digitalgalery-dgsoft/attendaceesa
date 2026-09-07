import openpyxl

excel_path = r'C:\Users\jamil\Downloads\Data Dulux\Stock 2026 (1)\Stock End 2026\202607 Stock End Juli 2026 AMK AKP Updt CKDB and ACE.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws_sub = wb['Submissions']

brands = set()
products = set()
colors = set()
keterangans = set()
galons = set()
pails = set()

for r in range(2, ws_sub.max_row + 1):
    brands.add(ws_sub.cell(r, 8).value)
    products.add(ws_sub.cell(r, 9).value)
    colors.add(ws_sub.cell(r, 10).value)
    keterangans.add(ws_sub.cell(r, 7).value)
    galons.add(ws_sub.cell(r, 11).value)
    pails.add(ws_sub.cell(r, 13).value)

print('Brands:', sorted([str(b) for b in brands if b is not None]))
print('\nProducts (count={}):'.format(len(products)), sorted([str(p) for p in products if p is not None]))
print('\nColors:', sorted([str(c) for c in colors if c is not None]))
print('\nKeterangan:', sorted([str(k) for k in keterangans if k is not None]))
print('\nKemasan Galon:', sorted([g for g in galons if g is not None]))
print('\nKemasan Pail:', sorted([p for p in pails if p is not None]))

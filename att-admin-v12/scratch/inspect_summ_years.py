import openpyxl

excel_path = r'C:\Users\jamil\Downloads\Data Dulux\Stock 2026 (1)\Stock End 2026\202607 Stock End Juli 2026 AMK AKP Updt CKDB and ACE.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws_summ = wb['Summ']

# Check row 1, 2, 3, 4, 5
for c in range(18, ws_summ.max_column + 1, 5):
    r1 = ws_summ.cell(1, c).value
    r3 = ws_summ.cell(3, c).value
    r4 = ws_summ.cell(4, c).value
    r5 = ws_summ.cell(5, c).value
    print(f'Col {c}: r1="{r1}", r3="{r3}", r4="{r4}", r5="{r5}"')

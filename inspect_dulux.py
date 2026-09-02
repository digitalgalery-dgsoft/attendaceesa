import openpyxl

path = r"C:\Users\jamil\OneDrive\Documents\Reporting Prinsiple\Dulux\Copy of Template Report Untuk Sadata.xlsx"
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
print("DULUX Sheets:", wb.sheetnames)
for sname in wb.sheetnames:
    ws = wb[sname]
    headers = []
    for row in ws.iter_rows(values_only=True):
        if any(row):
            headers = [str(c).strip() for c in row if c is not None]
            break
    print(f"\n[Dulux Sheet: {sname}] (Count: {len(headers)})")
    print(f"   Fields: {', '.join(headers)}")
